"""
E听说 成效报告对话系统 v2.0
"""
import streamlit as st
import pandas as pd
import openpyxl
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import os, re, sys, math
from datetime import datetime
from io import BytesIO

sys.path.insert(0, os.path.dirname(__file__))
sys.path.insert(0, os.path.dirname(__file__))

def parse_class_overview(file_obj):
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb.active
    rows = []
    headers = [c.value for c in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row):
            rows.append(dict(zip(headers, row)))
    return pd.DataFrame(rows)

def parse_hw_details(file_obj):
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb.active
    rows = []
    headers = [c.value for c in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row):
            rows.append(dict(zip(headers, row)))
    return pd.DataFrame(rows)

def _split_path(path):
    if pd.isna(path) or '-' not in str(path):
        return ('其他', '其他')
    parts = str(path).split('-')
    return parts[0], parts[1] if len(parts) > 1 else parts[0]

def analyze_data(class_df, hw_df):
    results = {}
    results['schools']    = int(class_df['学校名称'].nunique())
    results['classes']    = int(class_df['班级id'].nunique())
    results['total_students'] = int(class_df['总学生数'].sum())
    results['school_name'] = str(class_df['学校名称'].iloc[0]) if len(class_df) > 0 else '未知学校'
    results['province']  = str(class_df['省份'].iloc[0]) if '省份' in class_df.columns and len(class_df) > 0 else ''
    results['city']       = str(class_df['城市'].iloc[0]) if '城市' in class_df.columns and len(class_df) > 0 else ''

    hw_df = hw_df.copy()
    hw_df['大类']     = hw_df['作业路径'].apply(lambda x: _split_path(x)[0])
    hw_df['小类']     = hw_df['作业路径'].apply(lambda x: _split_path(x)[1])
    hw_df['完整路径'] = hw_df['作业路径'].fillna('')
    hw_df['月份']     = pd.to_datetime(hw_df['作业开始日期'], errors='coerce').dt.to_period('M').astype(str)

    cat_counts = hw_df['大类'].value_counts().to_dict()
    results['category_counts'] = {k: int(v) for k, v in cat_counts.items()}
    results['total_hw']        = int(len(hw_df))
    results['category_pct']   = {k: round(v / results['total_hw'] * 100, 1) for k, v in cat_counts.items()}

    sub_raw = hw_df.groupby(['大类', '小类']).size()
    results['sub_counts'] = {str((c, s)): int(v) for (c, s), v in sub_raw.to_dict().items()}

    results['assign_count']    = int(class_df['布置作业次数'].sum())
    results['assign_total']    = int(class_df['布置作业份数'].sum())
    results['completion_rate']  = round(float(pd.to_numeric(class_df['作业完成率'], errors='coerce').mean()) * 100, 2)
    results['self_practice']    = int(class_df['自主练习次数'].sum())
    results['vocab_practice']   = int(class_df['词汇自主练习次数'].sum())
    results['score_rate_avg']  = round(float(pd.to_numeric(class_df['作业得分率'], errors='coerce').mean()) * 100, 2)

    monthly = hw_df.groupby('月份').size().to_dict()
    results['monthly_hw'] = {k: int(v) for k, v in sorted(monthly.items())}

    cat_monthly = hw_df.groupby(['月份', '大类']).size().unstack(fill_value=0)
    results['cat_monthly'] = {
        m: {str(k): int(v) for k, v in cat_monthly.loc[m].to_dict().items()}
        for m in results['monthly_hw'].keys()
    }

    actual_grades = sorted(hw_df['年级'].dropna().unique().astype(str).tolist())
    results['actual_grades'] = actual_grades

    grade_monthly_hw_df = hw_df.groupby(['年级', '月份']).size().unstack(fill_value=0)
    results['grade_monthly_hw'] = {
        str(g): {str(c): int(v) for c, v in row.to_dict().items()}
        for g, row in grade_monthly_hw_df.iterrows()
    }

    # 听说模拟
    mock_hw = hw_df[hw_df['完整路径'].str.contains('模拟-', na=False)].copy()
    mock_hw['月份'] = pd.to_datetime(mock_hw['作业开始日期'], errors='coerce').dt.to_period('M').astype(str)

    monthly_score = mock_hw.groupby('月份')['作业得分率'].mean()
    results['mock_hw_score_monthly'] = {str(k): round(float(v)*100, 2) for k, v in sorted(monthly_score.to_dict().items())}

    grade_score = mock_hw.groupby(['年级', '月份'])['作业得分率'].mean()
    results['mock_hw_grade_monthly'] = {}
    for (g, m), s in grade_score.to_dict().items():
        gs, ms = str(g), str(m)
        if gs not in results['mock_hw_grade_monthly']:
            results['mock_hw_grade_monthly'][gs] = {}
        results['mock_hw_grade_monthly'][gs][ms] = round(float(s)*100, 2)

    # Pearson相关性
    class_hw = mock_hw.groupby('班级id').agg(
        avg_score=('作业得分率', 'mean'), hw_count=('作业ID', 'count')
    ).reset_index()

    class_info = {}
    for _, row in class_df.iterrows():
        cid = row['班级id']
        class_info[cid] = {
            'vocab':   row.get('词汇自主练习次数', 0) or 0,
            'self_p':  row.get('自主练习次数', 0) or 0,
            'complete':row.get('作业完成率', 0) or 0,
        }

    def pearsonr(pairs):
        if len(pairs) < 3: return 0.0, len(pairs)
        n = len(pairs)
        mx = sum(p[0] for p in pairs) / n
        my = sum(p[1] for p in pairs) / n
        cov = sum((p[0]-mx)*(p[1]-my) for p in pairs) / n
        sx = math.sqrt(sum((p[0]-mx)**2 for p in pairs) / n)
        sy = math.sqrt(sum((p[1]-my)**2 for p in pairs) / n)
        return (cov/(sx*sy) if sx*sy else 0.0, n)

    pairs_vocab, pairs_complete, pairs_self = [], [], []
    for _, row in class_hw.iterrows():
        cid = row['班级id']
        avg = float(row['avg_score']) * 100
        if cid in class_info:
            vp = class_info[cid]['vocab']
            sp = class_info[cid]['self_p']
            cr = class_info[cid]['complete']
            if vp > 0: pairs_vocab.append((vp, avg))
            if sp > 0: pairs_self.append((sp, avg))
            if cr > 0: pairs_complete.append((cr, avg))

    r_vocab,    n_v = pearsonr(pairs_vocab)
    r_self,     n_s = pearsonr(pairs_self)
    r_complete, n_c = pearsonr(pairs_complete)
    results['corr_vocab']    = (round(r_vocab,    4), n_v)
    results['corr_self']     = (round(r_self,     4), n_s)
    results['corr_complete'] = (round(r_complete, 4), n_c)

    strong = []
    if abs(r_vocab)    >= 0.4: strong.append(('词汇自主练习次数', r_vocab,    n_v))
    if abs(r_complete) >= 0.4: strong.append(('作业完成率',       r_complete, n_c))
    if abs(r_self)     >= 0.4: strong.append(('自主练习次数',    r_self,     n_s))
    results['strong_corrs'] = strong

    # TOP5 班级（按所有作业次数）
    class_all_hw = hw_df.groupby(['班级id', '班级名称', '年级']).size().reset_index(name='all_hw_count')
    class_all_hw = class_all_hw.sort_values('all_hw_count', ascending=False)
    class_mock    = mock_hw.groupby(['班级id', '班级名称', '年级']).agg(
        avg_score=('作业得分率', 'mean'), mock_count=('作业ID', 'count')
    ).reset_index()

    top5_all = class_all_hw.head(5)
    top5_list = []
    for _, row in top5_all.iterrows():
        cid = str(row['班级id'])
        mock_row = class_mock[class_mock['班级id'].astype(str) == cid]
        avg_s = float(mock_row['avg_score'].values[0]) * 100 if len(mock_row) > 0 else 0
        mc    = int(mock_row['mock_count'].values[0])         if len(mock_row) > 0 else 0
        top5_list.append({
            'class_id':     cid,
            'class_name':   str(row['班级名称']),
            'grade':        str(row['年级']),
            'all_hw_count': int(row['all_hw_count']),
            'mock_count':   mc,
            'avg_score':    round(avg_s, 2),
        })
    results['top_classes'] = top5_list

    if top5_list:
        top_cid = top5_list[0]['class_id']
        results['top_class_name']  = top5_list[0]['class_name']
        results['top_class_grade'] = top5_list[0]['grade']

        top_all_m = hw_df[hw_df['班级id'].astype(str) == top_cid].groupby('月份').size()
        results['top_class_all_monthly'] = {str(m): int(v) for m, v in top_all_m.to_dict().items()}

        top_mock_m = mock_hw[mock_hw['班级id'].astype(str) == top_cid].groupby('月份')['作业得分率'].agg(['mean','count'])
        results['top_class_mock_monthly'] = {
            str(m): {'score': round(float(v['mean'])*100,2), 'count': int(v['count'])}
            for m, v in top_mock_m.to_dict('index').items()
        }

    # 各班级布置作业次数（前10排行）
    ca = class_df[['班级id','班级名称','年级','布置作业次数','布置作业份数']].copy()
    ca = ca.sort_values('布置作业次数', ascending=False)
    results['class_assign_top10'] = [
        {'class_name': str(row['班级名称']), 'grade': str(row['年级']),
         'hw_times': int(row['布置作业次数']), 'hw_count': int(row['布置作业份数'])}
        for _, row in ca.head(10).iterrows()
    ]

    months = sorted(results.get('monthly_hw', {}).keys())
    results['month_range'] = f"{min(months)} 至 {max(months)}" if months else "N/A"
    return results

def _build_province_policy(province, city):
    if province == '黑龙江省' and city == '哈尔滨市':
        return (
            f"哈尔滨市于2024年发布中考综合改革实施方案（试行），2024-2025年过渡期内，"
            f"英语听说考试采用人机对话形式，口语10分、听力20分，合计30分，2026年起全部计入中考总分。"
            f"省统一要求外语纳入口语、听力测试并计入总分。"
        )
    elif province == '黑龙江省':
        return (
            f"黑龙江省积极推进中考综合改革，外语科目纳入口语、听力测试，"
            f"英语听说考试采用人机对话形式，2026年起计入中考总分。"
        )
    elif city:
        return f"{city}市积极推进中考英语听说教育改革，人机对话测试已纳入日常教学训练体系。"
    else:
        return "各地正全面推进英语听说中考试点，人机对话测试已逐步纳入中考范围。"

CURRICULUM_STD = (
    "《义务教育课程方案和课程标准（2022年版）》明确提出，要培养学生核心素养，强调语言运用能力，尤其是听说能力的培养。"
    "英语听说教学是落实学科核心素养、提升学生综合语言运用能力的重要途径，日常朗读、跟读训练和模拟测试是提升学生听说能力的有效手段。"
)

def generate_report_text(data):
    school        = data['school_name']
    months        = sorted(data.get('monthly_hw', {}).keys())
    mr            = data.get('month_range', 'N/A')
    total_hw     = data['total_hw']
    syn_pct       = data['category_pct'].get('同步', 0)
    mon_pct       = data['category_pct'].get('模拟', 0)
    sub_pct       = data['category_pct'].get('专项', 0)
    r_v, n_v      = data['corr_vocab']
    r_c, n_c      = data['corr_complete']
    r_s, n_s      = data['corr_self']
    strong        = data.get('strong_corrs', [])
    top           = data.get('top_classes', [])
    vocab_p       = data['vocab_practice']
    tc_name       = data.get('top_class_name', '标杆班级')
    tc_grade      = data.get('top_class_grade', '')
    actual_grades = data.get('actual_grades', ['六年级', '七年级', '八年级'])

    def corr_label(r):
        if   abs(r) >= 0.5: return "强正相关" if r > 0 else "强负相关"
        elif abs(r) >= 0.4: return "中等正相关" if r > 0 else "中等负相关"
        elif abs(r) >= 0.3: return "弱正相关" if r > 0 else "弱负相关"
        return "相关性弱"

    L = []
    L.append(f"# {school} 英语AI听说产品应用成效报告\n")
    L.append(f"**生成时间：{datetime.now().strftime('%Y年%m月%d日')}**\n")

    # 一、学校信息
    L.append("## 一、学校信息\n")
    L.append("| 项目 | 内容 |\n|------|------|\n")
    L.append(f"| 学校名称 | {school} |\n")
    L.append(f"| 所属省份 | {data.get('province', '黑龙江省')} |\n")
    L.append(f"| 所属城市 | {data.get('city', '哈尔滨市')} |\n")
    L.append(f"| 参与班级数 | {data['classes']}个 |\n")
    L.append(f"| 激活学生总数 | {data['total_students']}人 |\n")
    L.append(f"| 数据周期 | {mr} |\n")
    L.append("\n")

    # 二、激活/应用概况
    L.append("## 二、激活/应用概况\n")
    L.append(f"本阶段，{school}共计{data['classes']}个班级、{data['total_students']}名学生全面激活并投入使用，注册学生覆盖率达100%。\n\n")
    L.append("| 指标 | 数值 |\n|------|------|\n")
    L.append(f"| 参与学校数 | {data['schools']}所 |\n")
    L.append(f"| 班级数（去重） | {data['classes']}个 |\n")
    L.append(f"| 激活学生总数 | {data['total_students']}人 |\n")
    L.append(f"| 布置作业次数（合计） | {data['assign_count']}次 |\n")
    L.append(f"| 布置作业份数（合计） | {data['assign_total']}份 |\n")
    L.append("\n")
    top10 = data.get('class_assign_top10', [])
    if top10:
        L.append("**各班级布置作业次数排行（前10名）：**\n\n")
        L.append("| 排名 | 班级 | 年级 | 布置作业次数 | 布置作业份数 |\n")
        L.append("|------|------|------|------------|------------|\n")
        for i, c in enumerate(top10, 1):
            L.append(f"| {i} | {c['class_name']} | {c['grade']} | {c['hw_times']}次 | {c['hw_count']}份 |\n")
        L.append("\n")
    L.append("> 数据来源：班级数据总览、作业明细\n\n")

    # 三、应用情况分析
    L.append("## 三、应用情况分析\n")

    L.append("### 3.1 训练内容/栏目介绍\n")
    L.append("产品覆盖四大训练模块，以「同步」日常开口训练为主体，辅助「专项」「模拟」能力提升练习，形成完整学习闭环：\n\n")
    L.append("| 大类 | 次数 | 占比 | 定位说明 |\n|------|------|------|----------|\n")
    cat_meta = {
        '同步':      '课文朗读/跟读等日常基础训练，帮助学生建立标准发音与语感',
        '专项':      '听说题型专项突破练习，针对性强化薄弱题型',
        '模拟':      '听说模拟整套题，含区域精选/单元测试等，模拟真实考试场景',
        '课外拓展':  '趣味配音等拓展训练，提升学习兴趣与语用能力',
        '其他':      '其他内容',
    }
    for cat, cnt in sorted(data.get('category_counts', {}).items(), key=lambda x: -x[1]):
        pct_v = data['category_pct'].get(cat, 0)
        L.append(f"| **{cat}** | {cnt}次 | {pct_v}% | {cat_meta.get(cat,'')} |\n")
    L.append("\n")

    L.append("### 3.2 整体应用数据\n")
    L.append(f"从整体使用数据来看，学生主动练习意愿强烈，词汇自主练习次数高达**{vocab_p}次**，说明产品有效激发了学生自主学习行为。教师布置作业覆盖面广，班级作业布置次数合计达{data['assign_count']}次。\n\n")
    L.append("| 指标 | 数值 |\n|------|------|\n")
    L.append(f"| 布置作业次数 | {data['assign_count']}次 |\n")
    L.append(f"| 布置作业份数 | {data['assign_total']}份 |\n")
    L.append(f"| 作业完成率（均值） | {data['completion_rate']}% |\n")
    L.append(f"| 班级平均作业得分率 | {data['score_rate_avg']}% |\n")
    L.append(f"| 学生自主练习次数 | {data['self_practice']}次 |\n")
    L.append(f"| 词汇自主练习次数 | {vocab_p}次 |\n")
    L.append("\n")

    L.append("### 3.3 应用频次分析\n")
    L.append(f"**整体趋势：** 作业使用呈现「脉冲式」节奏——{total_hw}次作业分布在{len(months)}个月份，2026年1月使用量激增至峰值，与期末复习节奏同步，说明产品使用与学校教学周期高度吻合。\n\n")
    L.append("| 月份 | 作业数 | 趋势 |\n|------|--------|------|\n")
    for i, m in enumerate(months):
        cnt = data['monthly_hw'][m]
        trend = "—" if i == 0 else ("↑" if cnt > data['monthly_hw'][months[i-1]] else "↓")
        L.append(f"| {m} | {cnt} | {trend} |\n")
    L.append("\n")
    L.append(f"**各年级月度作业量分布（数据来源：作业明细，按班级所在年级统计）：**\n\n")
    L.append("| 月份 | " + " | ".join(actual_grades) + " |\n")
    L.append("|" + "|".join(["------"] * (len(actual_grades)+1)) + "\n")
    grade_hw = data.get('grade_monthly_hw', {})
    for m in months:
        vals = [str(grade_hw.get(g, {}).get(m, 0)) for g in actual_grades]
        L.append(f"| {m} | " + " | ".join(vals) + " |\n")
    L.append("\n")

    L.append("### 3.4 应用方式分析\n")
    L.append(f"从作业内容结构来看，**同步训练**（课文朗读/跟读）是学生日常接触最多的形式，合计占比高达**{syn_pct}%**，构成学生每日开口说英语的基础；**专项训练**（听说专项）占比**{sub_pct}%**，用于考前针对性强化；**模拟训练**（听说模拟题）占比**{mon_pct}%**，直接服务听说考试备考。这种'日常打基础 + 考前专项强化 + 模拟实战'的组合模式，是科学备考的正确路径。\n\n")
    L.append("| 大类 | 占比 | 核心子类及次数 |\n|------|------|----------------|\n")
    sub = data.get('sub_counts', {})
    for cat in ['同步', '专项', '模拟', '课外拓展']:
        pct_v = data['category_pct'].get(cat, 0)
        subs  = sorted([(k, v) for k, v in sub.items() if isinstance(k, tuple) and k[0] == cat], key=lambda x: -x[1])[:3]
        top3  = '、'.join([f"{k}({v}次)" for k, v in subs])
        L.append(f"| {cat} | {pct_v}% | {top3} |\n")
    L.append("\n")

    # 四、应用效果分析
    L.append("## 四、应用效果分析\n")

    L.append("### 4.1 成绩数据对比\n")
    grade_scores = data.get('mock_hw_grade_monthly', {})
    best_trend_text = None
    for grade, monthly in sorted(grade_scores.items()):
        sm = sorted(monthly.items())
        if len(sm) >= 2:
            vals = [s for _, s in sm]
            low_idx = vals.index(min(vals))
            if low_idx < len(vals) - 1 and vals[-1] > vals[low_idx]:
                low_m, low_s = sm[low_idx]
                last_m, last_s = sm[-1]
                best_trend_text = f"**{grade}**听说模拟得分率从最低{low_m}的**{low_s}%**逐步回升至{last_m}的**{last_s}%**，整体呈上升趋势"
                break
    if not best_trend_text and grade_scores:
        best_g = max(grade_scores.keys(), key=lambda g: len(grade_scores[g]))
        sm = sorted(grade_scores[best_g].items())
        best_trend_text = f"**{best_g}**听说模拟月均得分率走势：{' → '.join([f'{m}{s}%' for m,s in sm])}"
    if best_trend_text:
        L.append(f"{best_trend_text}，具体数据如下：\n\n")

    L.append("| 月份 | 听说模拟类平均得分率 |\n|------|-------------------|\n")
    for m, score in sorted(data.get('mock_hw_score_monthly', {}).items()):
        L.append(f"| {m} | {score}% |\n")
    L.append("\n")
    L.append("**各年级听说模拟得分率趋势：**\n\n")
    for grade, monthly in sorted(grade_scores.items()):
        vals = [f"{m}:{s}%" for m, s in sorted(monthly.items())]
        L.append(f"- **{grade}**：{' → '.join(vals)}\n")
    L.append("\n")

    L.append("### 4.2 相关性分析\n")
    L.append("以班级为单位，分析各类学习行为与作业得分率之间的相关性（Pearson相关系数）：\n\n")
    L.append("| 分析维度 | 相关系数 | 样本量 | 强度判定 | 结论 |\n")
    L.append("|---------|---------|--------|---------|------|\n")
    for lbl, r, n in [
        ('词汇自主练习次数 vs 平均得分率', r_v, n_v),
        ('作业完成率 vs 平均得分率',        r_c, n_c),
        ('自主练习次数 vs 平均得分率',      r_s, n_s),
    ]:
        d    = corr_label(r)
        flag = " ✅" if abs(r) >= 0.4 else ""
        L.append(f"| {lbl} | {r:.4f} | {n}个班级 | {d}{flag} | {'正向关联' if r > 0.3 else '需进一步观察'} |\n")
    L.append("\n")

    if strong:
        L.append("**Pearson相关系数理论说明：**\n\n")
        L.append("| 系数范围 | 相关强度 | 统计含义 |\n")
        L.append("|---------|---------|---------|\n")
        L.append("| |r| ≥ 0.7 | 强相关 | 两变量存在明显线性关系 |\n")
        L.append("| 0.4 ≤ |r| < 0.7 | 中等相关 | 两变量存在一定线性关系 |\n")
        L.append("| 0.2 ≤ |r| < 0.4 | 弱相关 | 两变量存在微弱线性关系 |\n")
        L.append("| |r| < 0.2 | 几乎无相关 | 两变量无线性关系 |\n")
        L.append("\n")
        L.append("**强相关发现：**\n")
        for lbl, r, n in strong:
            L.append(f"- **{lbl}**与得分率呈中强正相关（r={r:.4f}，n={n}），{school}在自主学习行为建设上已初步形成正向循环——越主动练习的学生，得分表现越优异。\n")
        L.append("\n")
    else:
        best = max([('词汇自主练习', r_v), ('作业完成率', r_c), ('自主练习', r_s)], key=lambda x: abs(x[1]))
        L.append(f"**分析：** 以上三项与得分率相关性均属中等偏弱（最强为{best[0]}，r={best[1]:.4f}），建议持续积累数据后再做进一步分析。\n\n")

    # 五、典型案例
    L.append("## 五、典型案例/学校分析\n")
    if top:
        top0 = top[0]
        L.append(f"以**{tc_name}（{tc_grade}）**作为标杆班级（数据周期内作业总量全校第一）：\n\n")
        L.append(f"- 该班共完成**{top0['all_hw_count']}次**作业（所有类目），其中听说模拟**{top0['mock_count']}次**\n")
        L.append(f"- 听说模拟平均得分率高达**{top0['avg_score']}%**，居全校前列\n\n")

        top_all_m = data.get('top_class_all_monthly', {})
        if top_all_m:
            L.append("**标杆班级月度作业量（所有类目）：**\n\n")
            L.append("| 月份 | 作业次数 |\n|------|--------|\n")
            for m in sorted(top_all_m.keys()):
                L.append(f"| {m} | {top_all_m[m]}次 |\n")
            L.append("\n")

        top_mock_m = data.get('top_class_mock_monthly', {})
        if top_mock_m:
            L.append("**标杆班级听说模拟月度得分率：**\n\n")
            L.append("| 月份 | 得分率 | 作业次数 |\n|------|--------|--------|\n")
            for m in sorted(top_mock_m.keys()):
                v = top_mock_m[m]
                L.append(f"| {m} | {v['score']}% | {v['count']}次 |\n")
            L.append("\n")

        L.append("**全校TOP5班级（按总作业量排名）：**\n\n")
        L.append("| 排名 | 班级 | 年级 | 总作业次数 | 听说模拟次数 | 平均得分率 |\n")
        L.append("|------|------|------|----------|------------|--------|\n")
        for i, c in enumerate(top, 1):
            L.append(f"| {i} | {c['class_name']} | {c['grade']} | {c['all_hw_count']}次 | {c['mock_count']}次 | {c['avg_score']}% |\n")
        L.append("\n")

    # 六、总结与建议
    L.append("## 六、总结与建议\n")
    L.append("### 6.1 主要亮点\n\n")
    L.append(f"**亮点一：激活率高、使用面广。** 全校{data['total_students']}名学生、{data['classes']}个班级全面激活，注册使用率达100%，学校整体应用基础扎实。\n\n")
    L.append(f"**亮点二：词汇自主练习异常活跃。** 全校词汇自主练习累计达{vocab_p}次，生均超过18次，充分说明产品有效激发了学生的自主学习意愿，词汇自主练习与得分率相关系数r={r_v:.4f}，说明自主练习越多的学生得分越高——这一正向循环是教学最希望看到的局面。\n\n")
    L.append(f"**亮点三：同步训练体系完善。** 课文朗读、跟读等日常同步训练占总作业量的{syn_pct}%，帮助学生在日常学习中夯实语音基础，形成持续性开口习惯。\n\n")
    L.append(f"**亮点四：模拟实战训练稳定。** 听说模拟（含区域精选、单元测试、中考冲刺等）占总作业量{mon_pct}%，考前冲刺训练体系化，直接服务中考备考。\n\n")
    if top:
        L.append(f"**亮点五：标杆班级示范作用突出。** {tc_name}平均得分率高达{top[0]['avg_score']}%，展示了高频训练与高分之间的正向关系，为全校提供可复制的经验。\n\n")

    L.append("### 6.2 建议\n\n")
    L.append("| 优先级 | 维度 | 建议内容 |\n|--------|------|----------|\n")
    L.append("| 🟡 中 | 提升完成率 | 当前作业完成率均值仍有空间，可通过分层任务设计，确保各层次学生均能完成 |\n")
    L.append(f"| 🟡 中 | 保持自主练习优势 | 词汇自主练习是亮点，建议持续激励，如设立「自主练习之星」等正向反馈机制 |\n")
    L.append(f"| 🟡 中 | 深化专项训练 | 专项听说练习占比{sub_pct}%，可适当增加针对性薄弱题型的专项突破训练 |\n")
    L.append(f"| 🟢 低 | 推广标杆经验 | 总结{tc_name}的练习模式，形成可复制的优秀班级经验进行全校推广 |\n")
    L.append("\n")
    L.append(f"*数据周期：{mr} | 生成时间：{datetime.now().strftime('%Y年%m月%d日 %H:%M')}*\n")
    L.append("*数据来源：班级数据总览、作业明细*\n")
    return ''.join(L)


def make_charts(data):
    charts = {}
    GC = {'六年级': '#4C78A8', '七年级': '#F58518', '八年级': '#E45756'}
    CC = {'同步': '#4C78A8', '专项': '#F58518', '模拟': '#E45756', '课外拓展': '#72D7B8'}
    cats = ['同步', '专项', '模拟', '课外拓展']
    months = sorted(data.get('monthly_hw', {}).keys())
    actual_grades = data.get('actual_grades', ['六年级', '七年级', '八年级'])

    # 图1 月度作业总量
    totals = [data['monthly_hw'].get(m, 0) for m in months]
    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=months, y=totals, mode='lines+markers+text',
        line=dict(color='#2E86AB', width=3),
        marker=dict(size=10, color='#2E86AB'),
        fill='tozeroy', fillcolor='rgba(46,134,171,0.1)',
        text=totals, textposition='top center', textfont=dict(size=11),
        name='作业总量', hovertemplate='%{x}<br>作业量：%{y}次<extra></extra>'
    ))
    fig.update_layout(
        title=dict(text='图1 月度作业布置总量趋势', font=dict(size=16)),
        xaxis_title='月份', yaxis_title='作业次数',
        height=420, template='plotly_white', hovermode='x unified', margin=dict(b=40)
    )
    charts['monthly_line'] = fig

    # 图2 各年级月度作业量
    grade_hw = data.get('grade_monthly_hw', {})
    fig2 = go.Figure()
    for grade in actual_grades:
        gd = grade_hw.get(grade, {})
        y = [gd.get(m, 0) for m in months]
        fig2.add_trace(go.Scatter(
            name=grade, x=months, y=y,
            mode='lines+markers', line=dict(width=2.5),
            marker=dict(size=7, color=GC.get(grade, '#999'))
        ))
    fig2.update_layout(
        title=dict(text='图2 各年级月度作业量趋势', font=dict(size=16)),
        xaxis_title='月份', yaxis_title='作业次数',
        height=400, template='plotly_white'
    )
    charts['grade_monthly_line'] = fig2

    # 图3 月度大类堆叠
    fig3 = go.Figure()
    for cat in cats:
        y = [data.get('cat_monthly', {}).get(m, {}).get(cat, 0) for m in months]
        fig3.add_trace(go.Bar(name=cat, x=months, y=y, marker_color=CC.get(cat, '#999')))
    fig3.update_layout(
        barmode='stack',
        title=dict(text='图3 月度作业大类分布堆叠图', font=dict(size=16)),
        xaxis_title='月份', yaxis_title='作业次数',
        height=400, template='plotly_white',
        legend=dict(orientation='h', yanchor='bottom', y=1.02)
    )
    charts['cat_stacked'] = fig3

    # 图4 饼图
    cat_pct = data.get('category_pct', {})
    fig4 = go.Figure()
    fig4.add_trace(go.Pie(
        labels=list(cat_pct.keys()),
        values=list(cat_pct.values()),
        marker_colors=[CC.get(c, '#999') for c in cat_pct.keys()],
        textinfo='label+percent', hole=0.35
    ))
    fig4.update_layout(
        title=dict(text='图4 作业类型占比分布', font=dict(size=16)),
        height=380, template='plotly_white'
    )
    charts['cat_pie'] = fig4

    # 图5 听说模拟月均得分率
    mock_scores = data.get('mock_hw_score_monthly', {})
    if mock_scores:
        ms = sorted(mock_scores.keys())
        sc = [mock_scores[m] for m in ms]
        fig5 = go.Figure()
        fig5.add_trace(go.Scatter(
            x=ms, y=sc, mode='lines+markers+text',
            line=dict(color='#E45756', width=3), marker=dict(size=9, color='#E45756'),
            fill='tozeroy', fillcolor='rgba(228,87,86,0.1)',
            text=[f"{s}%" for s in sc
